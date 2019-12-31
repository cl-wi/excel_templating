import openpyxl
import openpyxl.drawing.image
import datetime
import subprocess
import sqlite3

LO = "C:\Program Files\LibreOffice\program\soffice.exe"
DB = "database.sqlite3"

# connect to database
conn = None
try:
    conn = sqlite3.connect(DB)
except Exception as e:
    print(e)

c = conn.cursor()


def data_1(cell):
    c.execute('SELECT "Data 1" FROM "table1" LIMIT 1')
    cell.value = c.fetchone()[0]


def data_2(cell):
    c.execute('SELECT "Data 2" FROM "table1" WHERE "Data 1" = 1 LIMIT 1')
    cell.value = c.fetchone()[0]


def data_3(cell):
    c.execute('SELECT "Data 3" FROM "table1" WHERE "Data 1" = 1 LIMIT 1')
    cell.value = c.fetchone()[0]


def name(cell):
    c.execute('SELECT "Name" FROM "table1" WHERE "Data 1" = 1 LIMIT 1')
    cell.value = c.fetchone()[0]


def config(cell):
    c.execute('SELECT "Config" FROM "table1" WHERE "Data 1" = 1 LIMIT 1')
    cell.value = c.fetchone()[0]


def comment(cell):
    c.execute('SELECT "Comment" FROM "table1" WHERE "Data 1" = 1 LIMIT 1')
    cell.value = c.fetchone()[0]


def timestamp(cell):
    c.execute('SELECT datetime("Timestamp") FROM "table1" WHERE "Data 1" = 1 LIMIT 1')
    cell.value = c.fetchone()[0]

def timestamp(cell):
    c.execute('SELECT datetime("Timestamp") FROM "table1" WHERE "Data 1"')
    cell.value = c.fetchone()[0]

def timestamp(cell):
    c.execute('SELECT datetime("Timestamp") FROM "table1" WHERE "Data 1" = 1 LIMIT 1')
    cell.value = c.fetchone()[0]

def timestamp_min(cell):
    c.execute('SELECT MIN(datetime("Timestamp")) FROM "table1" WHERE "Data 1" = 1')
    cell.value = c.fetchone()[0]

def timestamp_max(cell):
    c.execute('SELECT MAX(datetime("Timestamp")) FROM "table1" WHERE "Data 1" = 1')
    cell.value = c.fetchone()[0]

def timestamp_delta(cell):
    c.execute('SELECT MAX(strftime("%M", "Timestamp")) - MIN(strftime("%M", "Timestamp")) FROM "table1" WHERE "Data 1" = 1')
    cell.value = c.fetchone()[0]


def logo_1(cell, logo='logo.jpg'):
    img = openpyxl.drawing.image.Image(logo)
    ws.add_image(img, cell.coordinate)


def date(cell):
    cell.value = datetime.datetime.now()


functions = {
    'date' : date,
    'data_1': data_1,
    'data_2': data_2,
    'data_3': data_3,
    'name': name,
    'config': config,
    'comment': comment,
    'logo_1': logo_1,
    'timestamp' : timestamp,
    'timestamp_min' : timestamp_min,
    'timestamp_max' : timestamp_max,
    'timestamp_delta' : timestamp_delta
}


wb = openpyxl.load_workbook('template.xlsx')
ws = wb.active

for row in ws.iter_rows(values_only=True):
    print (row)

# search for '$' strings and try to replace
for row in ws.iter_rows():
    for cell in row:
        if type(cell.value) is str:
            if cell.value[:1] == '$':
                print("replacing {}".format(cell.value))
                if cell.value[1:] not in functions:
                    print ("!!!WARNING!!! {} not implemented".format(cell.value[1:]))
                try:
                    functions[cell.value[1:]](cell)
                except Exception as e:
                    print(e)

wb.save("protocol.xlsx")

conn.close()

# convert to pdf
subprocess.call("{} --headless --convert-to pdf protocol.xlsx".format(LO))